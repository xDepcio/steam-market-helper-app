from seleniumwire import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from config import Config as cfg
from time import sleep
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from selenium.common.exceptions import ElementNotInteractableException
import requests
import openpyxl
import json
from seleniumwire.utils import decode
from utils import scrape_reqs_for_inv_cards, scrape_reqs_for_graphs


class PostSellApp:
    def __init__(self):
        chrome_options = Options()
        chrome_options.add_experimental_option("detach", True)
        driver = webdriver.Chrome(options=chrome_options)
        driver.get("https://steamcommunity.com/login/home/?goto=profiles%2F76561198462126877")

        self.driver = driver
        self.offset = 0

    def login(self):
        login_field = self._look_for_ele(By.XPATH, '/html/body/div[1]/div[7]/div[4]/div[1]/div[1]/div/div/div/div[2]/div/form/div[1]/input')
        login_field.send_keys(cfg.LOGIN)
        pass_field = self._look_for_ele(By.XPATH, '/html/body/div[1]/div[7]/div[4]/div[1]/div[1]/div/div/div/div[2]/div/form/div[2]/input')
        pass_field.send_keys(cfg.PASSWORD)
        login_btn = self._look_for_ele(By.XPATH, '/html/body/div[1]/div[7]/div[4]/div[1]/div[1]/div/div/div/div[2]/div/form/div[4]/button')
        login_btn.click()
        eq_btn = self._look_for_ele(By.XPATH, '/html/body/div[1]/div[7]/div[6]/div[1]/div[3]/div/div[1]/div[2]/div[3]/div[2]/a/span[1]')
        self.driver.get('https://steamcommunity.com/profiles/76561198462126877/inventory/#753')

    def get_cards(self, sell_prices):
        card_eles = self.driver.find_elements(By.CLASS_NAME, 'itemHolder')
        for ele in card_eles:
            try:
                ele.click()
                sleep(1)
            except ElementNotInteractableException:
                self.driver.execute_script('return InventoryNextPage()')
                sleep(1)
                ele.click()
            card_name = self.driver.execute_script('return g_ActiveInventory.selectedItem.description.name')
            if card_name in sell_prices:
                asset = self.driver.execute_script('return g_ActiveInventory.selectedItem.assetid')
                print('IS IN', asset)
                self.post_sell_request(1)
            else:
                print('IS NOT IN')
        # print(card_eles)

    def post_sell_request(self, price):
        asset_id = self.driver.execute_script('return g_ActiveInventory.selectedItem.assetid')
        app_id = self.driver.execute_script('return g_ActiveInventory.selectedItem.appid')
        context_id = self.driver.execute_script('return g_ActiveInventory.selectedItem.contextid')
        session_id = self.driver.execute_script('return g_sessionID')
        connected_body_str = f"sessionid={session_id}&appid={app_id}&contextid={context_id}&assetid={asset_id}&amount=1&price=88"

        # promise = self.driver.execute_async_script("""
        #     fetch("https://steamcommunity.com/market/sellitem/", {
        #     "headers": {
        #     "content-type": "application/x-www-form-urlencoded; charset=UTF-8"
        #     }
        #     "body": "dfasd=dasd"
        #     "method": "POST"
        #     });
        #     """)
        promise = self.driver.execute_async_script("""
            return new Promise((resolve, reject) => {
                setTimeout(() => {
                    resolve('Hello World!');
                }, 1000);
            });
        """)


        async def handle_promise(promise):
            result = await promise
            print(result)

        handle_promise(promise)


        #     headers={
        #     "content-type": "application/x-www-form-urlencoded; charset=UTF-8",
        # }, url="https://steamcommunity.com/market/sellitem/", data=connected_body_str)

        # print(response.status_code)


    def get_sell_prices(self, path):
        workbook = openpyxl.load_workbook(path)
        worksheet = workbook.active
        cards_prices = {}

        while True:
            card_name = worksheet.cell(2+self.offset, 1).value
            sell_price = worksheet.cell(10+self.offset, 10).value
            if card_name is None:
                break
            cards_prices[card_name] = sell_price
            self.offset += 11

        return cards_prices


    def _look_for_ele(self, by, value):
        ele = None
        try:
            ele = self.driver.find_element(by, value)
        except Exception:
            sleep(0.2)
            ele = self._look_for_ele(by, value)
        return ele

    def get_price_data_by_item_nameid(self, itemname_id):
        sell, buy = scrape_reqs_for_graphs(self.driver)
        print(sell)
        print(buy)

    def get_itemname_id(self, item_url):
        """Returns itemname_id"""
        self.driver.get(item_url)
        itemname_id = self.driver.execute_script("return ItemActivityTicker.m_llItemNameID")
        return itemname_id

    def get_card_name(self):
        """On card page returns card name"""
        card_name_ele = self.driver.find_element(By.ID, 'largeiteminfo_item_name')
        card_name = card_name_ele.text
        return card_name

    def get_game_name(self):
        """On card page returns game name"""
        game_ele = self.driver.find_element(By.ID, 'largeiteminfo_item_type')
        game_name = game_ele.text
        return game_name

    def draw_data_for_item(self, path, item_url):
        item_id = self.get_itemname_id(item_url)
        sleep(0.5)
        try:
            sell, buy = scrape_reqs_for_graphs(self.driver, item_id)
        except Exception:
            return
        card_name = self.get_card_name()
        game_name = self.get_game_name()
        # card_volume = self.get_card_volume()
        card_volume = ''

        workbook = openpyxl.load_workbook(path)
        worksheet = workbook.active

        self._draw_static_data(worksheet, card_name, game_name, card_volume)
        self._draw_fetched_prices(worksheet, sell, buy)
        self._draw_margins(worksheet, sell, buy, 0, 0)
        self._draw_margins(worksheet, sell, buy, 5, 1)
        self._draw_margins(worksheet, sell, buy, 10, 2)
        self.offset += 11

        workbook.save(path)

    def _draw_margins(self, worksheet, sell, buy, ignore_num, n_entry):
        for price, amount in buy:
            if amount > ignore_num:
                highest_buy = price
                break

        for price, amount in sell:
            if amount > ignore_num:
                lowest_sell = price
                break

        sold_for = lowest_sell - 0.01
        bought_for = highest_buy + 0.01

        fees_dict = self.driver.execute_script(f"return CalculateFeeAmount({sold_for*100}, 0.1)")
        for_me = fees_dict['amount']/100 - fees_dict['fees']/100

        profit = for_me - bought_for


        worksheet.cell(5+n_entry+self.offset, 7).value = sold_for
        worksheet.cell(5+n_entry+self.offset, 8).value = for_me
        worksheet.cell(5+n_entry+self.offset, 9).value = bought_for
        worksheet.cell(5+n_entry+self.offset, 10).value = profit

    def _draw_fetched_prices(self, worksheet, sell, buy):
        i = 4
        for price, count in sell[:7]:
            print(price, count)
            worksheet.cell(i+self.offset, 1).value = price
            worksheet.cell(i+self.offset, 2).value = count
            i += 1

        i = 4
        for price, count in buy[:7]:
            print(price, count)
            worksheet.cell(i+self.offset, 4).value = price
            worksheet.cell(i+self.offset, 5).value = count
            i += 1

    def _draw_static_data(self, worksheet, card_name, game_name, card_volume):
        """Draw headers, merge cells..."""
        worksheet.cell(1+self.offset, 1).value = game_name
        worksheet.cell(2+self.offset, 1).value = card_name
        worksheet.merge_cells(start_row=1+self.offset, start_column=1, end_row=1+self.offset, end_column=10)
        worksheet.merge_cells(start_row=2+self.offset, start_column=1, end_row=2+self.offset, end_column=10)

        worksheet.cell(3+self.offset, 1).value = "Sell offers"
        worksheet.cell(3+self.offset, 4).value = "Buy offers"
        worksheet.merge_cells(start_row=3+self.offset, start_column=1, end_row=3+self.offset, end_column=2)
        worksheet.merge_cells(start_row=3+self.offset, start_column=4, end_row=3+self.offset, end_column=5)

        worksheet.cell(3+self.offset, 7).value = "profit margins"
        worksheet.cell(4+self.offset, 7).value = "sold for"
        worksheet.cell(4+self.offset, 8).value = "sold for (no fees)"
        worksheet.cell(4+self.offset, 9).value = "bought for"
        worksheet.cell(4+self.offset, 10).value = "margin"
        worksheet.merge_cells(start_row=3+self.offset, start_column=7, end_row=3+self.offset, end_column=10)

        worksheet.cell(10+self.offset, 7).value = 'Volume:'
        worksheet.cell(10+self.offset, 8).value = card_volume
        worksheet.cell(10+self.offset, 9).value = 'Sprzedaj za:'
        worksheet.merge_cells(start_row=11+self.offset, start_column=1, end_row=11+self.offset, end_column=10)


    def get_card_volume(self):
        """On card page get card 24h volume"""
        url = self.driver.current_url
        market_hash_name = url.split('/')[-1]
        data = requests.get(
            f"https://steamcommunity.com/market/priceoverview/?country=PL&currency=6&appid=753&market_hash_name={market_hash_name}"
        ).json()
        sleep(0.5)
        return data['volume']


def main():
    app = PostSellApp()
    app.login()

    sell_prices = app.get_sell_prices('inv_data.xlsx')
    print(sell_prices)

    app.get_cards(sell_prices)



if __name__ == '__main__':
    main()
