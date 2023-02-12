from seleniumwire import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from config import Config as cfg
from time import sleep
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
import requests
import openpyxl
import json
from seleniumwire.utils import decode
from utils import scrape_reqs_for_graphs, load_urls


class App:
    def __init__(self):
        chrome_options = Options()
        chrome_options.add_experimental_option("detach", True)
        driver = webdriver.Chrome(options=chrome_options)
        driver.get("https://steamcommunity.com/login/home/?goto=profiles%2F76561198462126877")

        self.driver = driver
        self.offset = 0
        # sleep(3)
        # self.get_price_data_by_item_nameid(1)
        # return

    def login(self):
        login_field = self._look_for_ele(By.XPATH, '/html/body/div[1]/div[7]/div[4]/div[1]/div[1]/div/div/div/div[2]/div/form/div[1]/input')
        login_field.send_keys(cfg.LOGIN)
        pass_field = self._look_for_ele(By.XPATH, '/html/body/div[1]/div[7]/div[4]/div[1]/div[1]/div/div/div/div[2]/div/form/div[2]/input')
        pass_field.send_keys(cfg.PASSWORD)
        login_btn = self._look_for_ele(By.XPATH, '/html/body/div[1]/div[7]/div[4]/div[1]/div[1]/div/div/div/div[2]/div/form/div[4]/button')
        login_btn.click()
        eq_btn = self._look_for_ele(By.XPATH, '/html/body/div[1]/div[7]/div[6]/div[1]/div[3]/div/div[1]/div[2]/div[3]/div[2]/a/span[1]')
        self.driver.get('https://steamcommunity.com/market/listings/753/1245620-Site%20of%20Lost%20Grace')

    def _look_for_ele(self, by, value):
        ele = None
        try:
            ele = self.driver.find_element(by, value)
        except Exception:
            sleep(0.2)
            ele = self._look_for_ele(by, value)
        return ele

    def get_price_data_by_item_nameid(self, itemname_id):
        sell, buy = scrape_reqs_for_graphs(self.driver)
        print(sell)
        print(buy)

    def get_itemname_id(self, item_url):
        """Returns itemname_id"""
        self.driver.get(item_url)
        itemname_id = self.driver.execute_script("return ItemActivityTicker.m_llItemNameID")
        return itemname_id

    def get_card_name(self):
        """On card page returns card name"""
        card_name_ele = self.driver.find_element(By.ID, 'largeiteminfo_item_name')
        card_name = card_name_ele.text
        return card_name

    def get_game_name(self):
        """On card page returns game name"""
        game_ele = self.driver.find_element(By.ID, 'largeiteminfo_item_type')
        game_name = game_ele.text
        return game_name

    def get_card_amount(self):
        """returns total sell orders on market"""
        try:
            total = self.driver.find_element(By.XPATH, '/html/body/div[1]/div[7]/div[2]/div[1]/div[4]/div/div[2]/div/div[4]/div[4]/div[1]/div/div[1]/div/span[1]').text
        except Exception:
            return 0
        return int(total)

    def draw_data_for_item(self, path, item_url):
        item_id = self.get_itemname_id(item_url)
        sleep(0.5)
        card_amount = self.get_card_amount()
        if card_amount == 0: return
        try:
            sell, buy = scrape_reqs_for_graphs(self.driver, item_id)
        except Exception:
            return
        card_name = self.get_card_name()
        game_name = self.get_game_name()
        # card_volume = self.get_card_volume()
        card_volume = ''

        workbook = openpyxl.load_workbook(path)
        worksheet = workbook.active

        self._draw_static_data(worksheet, card_name, game_name, card_volume, card_amount, item_url)
        self._draw_fetched_prices(worksheet, sell, buy)
        self._draw_margins(worksheet, sell, buy, 0, 0)
        self._draw_margins(worksheet, sell, buy, 5, 1)
        profit = self._draw_margins(worksheet, sell, buy, 10, 2)

        print(profit, profit > 0.01)
        print(card_amount, card_amount > 200)
        if profit >= 0.02 and card_amount > 500:
            print('YESSSSIR')
            self.offset += 12

        workbook.save('data.xlsx')

    def _draw_margins(self, worksheet, sell, buy, ignore_num, n_entry):
        highest_buy = 0
        for price, amount in buy:
            if amount > ignore_num:
                highest_buy = price
                break

        lowest_sell = 0
        for price, amount in sell:
            if amount > ignore_num:
                lowest_sell = price
                break

        sold_for = lowest_sell - 0.01
        bought_for = highest_buy + 0.01

        fees_dict = self.driver.execute_script(f"return CalculateFeeAmount({sold_for*100}, 0.1)")
        for_me = fees_dict['amount']/100 - fees_dict['fees']/100

        profit = for_me - bought_for


        worksheet.cell(5+n_entry+self.offset, 7).value = sold_for
        worksheet.cell(5+n_entry+self.offset, 8).value = for_me
        worksheet.cell(5+n_entry+self.offset, 9).value = bought_for
        worksheet.cell(5+n_entry+self.offset, 10).value = profit

        print('T', profit, float(profit))
        return float(profit)

    def _draw_fetched_prices(self, worksheet, sell, buy):
        i = 4
        for price, count in sell[:7]:
            print(price, count)
            worksheet.cell(i+self.offset, 1).value = price
            worksheet.cell(i+self.offset, 2).value = count
            i += 1

        i = 4
        for price, count in buy[:7]:
            print(price, count)
            worksheet.cell(i+self.offset, 4).value = price
            worksheet.cell(i+self.offset, 5).value = count
            i += 1

    def _draw_static_data(self, worksheet, card_name, game_name, card_volume, card_amount, card_url):
        """Draw headers, merge cells..."""
        worksheet.cell(1+self.offset, 1).value = game_name
        worksheet.cell(2+self.offset, 1).value = card_name
        worksheet.merge_cells(start_row=1+self.offset, start_column=1, end_row=1+self.offset, end_column=10)
        worksheet.merge_cells(start_row=2+self.offset, start_column=1, end_row=2+self.offset, end_column=10)

        worksheet.cell(3+self.offset, 1).value = "Sell offers"
        worksheet.cell(3+self.offset, 4).value = "Buy offers"
        worksheet.merge_cells(start_row=3+self.offset, start_column=1, end_row=3+self.offset, end_column=2)
        worksheet.merge_cells(start_row=3+self.offset, start_column=4, end_row=3+self.offset, end_column=5)

        worksheet.cell(3+self.offset, 7).value = "profit margins"
        worksheet.cell(4+self.offset, 7).value = "sold for"
        worksheet.cell(4+self.offset, 8).value = "sold for (no fees)"
        worksheet.cell(4+self.offset, 9).value = "bought for"
        worksheet.cell(4+self.offset, 10).value = "margin"
        worksheet.merge_cells(start_row=3+self.offset, start_column=7, end_row=3+self.offset, end_column=10)

        worksheet.cell(10+self.offset, 7).value = 'Volume:'
        worksheet.cell(10+self.offset, 8).value = card_volume
        worksheet.merge_cells(start_row=11+self.offset, start_column=1, end_row=11+self.offset, end_column=10)

        worksheet.cell(10+self.offset, 9).value = 'Total sell offers:'
        worksheet.cell(10+self.offset, 10).value = card_amount
        worksheet.cell(11+self.offset, 1).value = card_url


    def get_card_volume(self):
        """On card page get card 24h volume"""
        url = self.driver.current_url
        market_hash_name = url.split('/')[-1]
        data = requests.get(
            f"https://steamcommunity.com/market/priceoverview/?country=PL&currency=6&appid=753&market_hash_name={market_hash_name}"
        ).json()
        sleep(0.5)
        return data['volume']


def main():
    app = App()
    app.login()

    urls = cfg.FOLLOWED_CARDS_URLS
    urls = load_urls('links.json')

    for url in urls:
        app.draw_data_for_item('data.xlsx', url)


if __name__ == '__main__':
    main()
